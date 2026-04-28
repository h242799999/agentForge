#!/usr/bin/env python3
"""
Shimano SDK xlsx → specs/ Markdown 转换器

支持两种来源：
  1.0.0  源：API接口式样书/（按模块分 xlsx）
  1.0.2  源：1.0.2/2.APIインターフェース仕様書/（全模块单文件）

用法：
  python3 scripts/extract-specs.py                     # 同时提取 1.0.0 + 1.0.2
  python3 scripts/extract-specs.py --version 1.0.0     # 仅提取 1.0.0
  python3 scripts/extract-specs.py --version 1.0.2     # 仅提取 1.0.2
  python3 scripts/extract-specs.py --sdk-root <path>   # 指定 SDK 根目录

输出目录结构：
  specs/
  ├── v1.0.0/
  │   ├── api-spec-Auth.md
  │   ├── api-spec-Connection.md
  │   └── ...
  └── v1.0.2/
      ├── api-spec-Auth.md
      ├── api-spec-Connection.md
      └── ...
"""

import sys
import os
import re
import argparse
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("错误：需要 openpyxl。运行：pip3 install openpyxl")
    sys.exit(1)


# ── 列索引常量（基于实际 xlsx 结构）────────────────────────
COL_LABEL = 2      # 字段标签（API定義 / 概要 / パラメータ 等）
COL_VALUE = 5      # 主要值
COL_VALUE2 = 6     # 副值（Exception 类名）
COL_TYPE = 8       # 参数类型
COL_DIRECTION = 11 # 参数方向（：In / ：Out）
COL_DESC = 13      # 参数描述 / ExceptionType
COL_EXC_DESC = 22  # Exception 描述

API_START_LABEL = "API定義"

SKIP_SHEETS = {
    "変更履歴", "API一覧", "WpsReserved_CellImgList",
    "CustomizeCategory関連性の星取表", "connect status",
    "Riding status", "Maintenance status",
}

CLASS_SHEET_KEYWORDS = ["クラス構成", "クラス構成 "]

# 1.0.2 单文件中 sheet 名 → 模块名映射（sheet 名即模块）
V102_MODULE_SHEETS = {
    "Connection", "Customize", "MyBike", "ShimanoLoader",
    "Auth", "User", "UserSetting", "Maintenance",
    "Activity", "FirmwareUpdate", "Riding",
}

# sheet 名 → 归属的逻辑模块（用于生成文件名）
SHEET_TO_MODULE = {
    "ShimanoLoader": "Auth",
    "ShimanoLoaderクラス構成 ": "Auth",
    "Auth": "Auth",
    "Auth クラス構成": "Auth",
    "User": "Auth",
    "Userクラス構成": "Auth",
    "UserSetting": "Setting",
    "UserSettingクラス構成": "Setting",
    "Connection": "Connection",
    "Connectionクラス構成": "Connection",
    "Unit": "Connection",
    "Customize": "Customize",
    "Customizeクラス構成": "Customize",
    "MyBike": "MyBike",
    "MyBikeクラス構成": "MyBike",
    "Maintenance": "Maintenance",
    "Maintenanceクラス構成": "Maintenance",
    "Activity": "Activity",
    "Activityクラス構成": "Activity",
    "FirmwareUpdate": "Update",
    "FirmwareUpdateクラス構成": "Update",
    "Riding": "Riding",
    "Riding クラス構成": "Riding",
    "GuestUser": "Auth",       # 1.0.2 新增
    "Exception": None,         # 跳过
}


# ── 基础解析工具 ────────────────────────────────────────────

def cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def get_row_values(row: tuple) -> dict:
    def get(idx):
        return cell_str(row[idx]) if idx < len(row) else ""
    return {
        "label":    get(COL_LABEL),
        "value":    get(COL_VALUE),
        "value2":   get(COL_VALUE2),
        "type":     get(COL_TYPE),
        "direction":get(COL_DIRECTION),
        "desc":     get(COL_DESC),
        "exc_desc": get(COL_EXC_DESC),
    }


def is_api_sheet(ws) -> bool:
    for row in ws.iter_rows(max_row=300, values_only=True):
        if len(row) > COL_LABEL and cell_str(row[COL_LABEL]) == API_START_LABEL:
            return True
    return False


def is_class_sheet(sheet_name: str) -> bool:
    return any(kw in sheet_name for kw in CLASS_SHEET_KEYWORDS)


# ── API 解析 ────────────────────────────────────────────────

def parse_api_sheet(ws) -> list[dict]:
    apis = []
    current = None
    in_exception = False
    in_params = False

    for row in ws.iter_rows(values_only=True):
        rv = get_row_values(row)
        label = rv["label"]

        if label == API_START_LABEL:
            if current:
                apis.append(current)
            current = {
                "definition": rv["value"],
                "id": "", "summary": "",
                "params": [], "exceptions": [],
                "returns": "", "visibility": "",
                "require_login": "", "require_online": "",
                "category": "",
            }
            in_exception = in_params = False

        elif current is None:
            continue

        elif label == "API ID":
            current["id"] = rv["value"]

        elif label == "概要":
            current["summary"] = rv["value"]

        elif label == "パラメータ":
            in_exception = False
            in_params = True
            if rv["value"] and rv["value"] != "なし":
                current["params"].append({
                    "name": rv["value"], "type": rv["type"],
                    "direction": rv["direction"].replace("：", "").strip(),
                    "desc": rv["desc"],
                })

        elif label == "Exception":
            in_exception = True
            in_params = False

        elif label == "戻り値":
            in_exception = in_params = False
            ret = rv["value"]
            if rv["value2"]:
                ret = f"{ret} {rv['value2']}"
            current["returns"] = ret

        elif label == "公開範囲":
            current["visibility"] = rv["value"]
            in_exception = in_params = False

        elif label == "SHIMANO IDログイン必須":
            current["require_login"] = rv["value"]

        elif label == "オンライン必須":
            current["require_online"] = rv["value"]

        elif label == "カテゴリ":
            current["category"] = rv["value"]

        else:
            if in_exception and rv["value2"] and rv["value2"] != "ExceptionClass":
                current["exceptions"].append({
                    "class": rv["value2"],
                    "type": rv["desc"],
                    "desc": rv["exc_desc"],
                })
            elif in_params and rv["value"] and label == "":
                current["params"].append({
                    "name": rv["value"], "type": rv["type"],
                    "direction": rv["direction"].replace("：", "").strip(),
                    "desc": rv["desc"],
                })

    if current:
        apis.append(current)
    return apis


def parse_class_sheet(ws) -> list[dict]:
    classes = []
    current_class = None

    for row in ws.iter_rows(values_only=True):
        rv = get_row_values(row)
        label = rv["label"]

        if label and re.search(r'の構成[：:]', label):
            if current_class:
                classes.append(current_class)
            current_class = {"name": label, "fields": []}

        elif label == "クラス定義":
            if current_class is None:
                current_class = {"name": rv["value"], "fields": []}
            else:
                current_class["definition"] = rv["value"]

        elif label == "説明" and current_class:
            current_class["description"] = rv["value"]

        elif current_class and rv["value"] and label == "":
            name = rv["value"]
            if name not in ("フィールド名", "Field Name", "名前"):
                current_class["fields"].append({
                    "name": name,
                    "type": rv["type"] or rv["value2"],
                    "desc": rv["desc"] or rv["exc_desc"],
                })

    if current_class:
        classes.append(current_class)
    return classes


# ── Markdown 生成 ───────────────────────────────────────────

def api_to_markdown(api: dict) -> str:
    lines = []
    api_id = f" _(ID: {api['id']})_" if api["id"] else ""
    lines.append(f"### `{api['definition']}`{api_id}")
    lines.append("")

    if api["summary"]:
        lines.append(f"**概要**：{api['summary']}")
        lines.append("")

    meta = []
    if api["visibility"]:      meta.append(f"公開範囲：{api['visibility']}")
    if api["require_login"]:   meta.append(f"SHIMANO IDログイン：{api['require_login']}")
    if api["require_online"]:  meta.append(f"オンライン：{api['require_online']}")
    if api["category"]:        meta.append(f"カテゴリ：{api['category']}")
    if meta:
        lines.append(f"_{' ｜ '.join(meta)}_")
        lines.append("")

    params = [p for p in api["params"] if p["name"] and p["name"] != "なし"]
    if params:
        lines += ["**パラメータ**", "",
                  "| 名前 | 型 | 方向 | 説明 |",
                  "|------|-----|------|------|"]
        for p in params:
            lines.append(f"| `{p['name']}` | `{p['type'] or '-'}` | {p['direction'] or 'In'} | {p['desc'] or '-'} |")
        lines.append("")
    else:
        lines += ["**パラメータ**：なし", ""]

    if api["returns"]:
        lines += [f"**戻り値**：`{api['returns']}`", ""]

    excs = [e for e in api["exceptions"] if e["class"]]
    if excs:
        lines += ["**Exception**", "",
                  "| ExceptionClass | ExceptionType | 概要 |",
                  "|---------------|--------------|------|"]
        for e in excs:
            lines.append(f"| `{e['class']}` | `{e['type']}` | {e['desc']} |")
        lines.append("")

    lines += ["---", ""]
    return "\n".join(lines)


def classes_to_markdown(classes: list[dict]) -> str:
    lines = []
    for cls in classes:
        lines.append(f"### {cls['name']}")
        lines.append("")
        if cls.get("description"):
            lines += [cls["description"], ""]
        if cls.get("fields"):
            lines += ["| フィールド | 型 | 説明 |", "|-----------|-----|------|"]
            for f in cls["fields"]:
                lines.append(f"| `{f['name']}` | `{f['type'] or '-'}` | {f['desc'] or '-'} |")
            lines.append("")
        lines += ["---", ""]
    return "\n".join(lines)


# ── 版本提取逻辑 ────────────────────────────────────────────

def process_xlsx_single_module(xlsx_path: Path, module: str,
                                output_dir: Path, version: str):
    """处理单模块 xlsx（1.0.0 模式：每个文件一个模块）"""
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

    md_lines = [
        f"# {module} モジュール API 仕様 (SDK {version})",
        "",
        f"> 自動抽出元：`{xlsx_path.name}`  ",
        f"> SDK バージョン：**{version}**  ",
        f"> 抽出日：{datetime.now().strftime('%Y-%m-%d')}",
        "",
        "---",
        "",
    ]

    api_count = class_count = 0

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        if not any(any(c is not None for c in row)
                   for row in ws.iter_rows(max_row=5, values_only=True)):
            continue

        if is_api_sheet(ws):
            apis = parse_api_sheet(ws)
            if apis:
                md_lines += [f"## {sheet_name}", ""]
                for api in apis:
                    md_lines.append(api_to_markdown(api))
                api_count += len(apis)

        elif is_class_sheet(sheet_name):
            classes = parse_class_sheet(ws)
            if classes:
                md_lines += [f"## {sheet_name}（クラス構成）", "",
                             classes_to_markdown(classes)]
                class_count += len(classes)

    out = output_dir / f"{module}.md"
    out.write_text("\n".join(md_lines), encoding="utf-8")
    print(f"    ✓ {api_count} 个 API，{class_count} 个类 → {out.relative_to(out.parent.parent.parent)}")
    return api_count


def process_xlsx_all_modules(xlsx_path: Path, output_dir: Path, version: str):
    """处理全模块单文件 xlsx（1.0.2 模式：一个文件含所有模块）"""
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

    # 按模块分组 sheet
    module_sheets: dict[str, list] = {}
    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        module = SHEET_TO_MODULE.get(sheet_name)
        if module is None:
            continue
        module_sheets.setdefault(module, []).append(sheet_name)

    total_apis = 0
    for module, sheets in sorted(module_sheets.items()):
        md_lines = [
            f"# {module} モジュール API 仕様 (SDK {version})",
            "",
            f"> 自動抽出元：`{xlsx_path.name}`（シート：{', '.join(sheets)}）  ",
            f"> SDK バージョン：**{version}**  ",
            f"> 抽出日：{datetime.now().strftime('%Y-%m-%d')}",
            "",
            "---",
            "",
        ]

        api_count = class_count = 0

        for sheet_name in sheets:
            ws = wb[sheet_name]
            if not any(any(c is not None for c in row)
                       for row in ws.iter_rows(max_row=5, values_only=True)):
                continue

            if is_api_sheet(ws):
                apis = parse_api_sheet(ws)
                if apis:
                    md_lines += [f"## {sheet_name}", ""]
                    for api in apis:
                        md_lines.append(api_to_markdown(api))
                    api_count += len(apis)

            elif is_class_sheet(sheet_name):
                classes = parse_class_sheet(ws)
                if classes:
                    md_lines += [f"## {sheet_name}（クラス構成）", "",
                                 classes_to_markdown(classes)]
                    class_count += len(classes)

        out = output_dir / f"{module}.md"
        out.write_text("\n".join(md_lines), encoding="utf-8")
        print(f"    ✓ {api_count} 个 API，{class_count} 个类 → {out.relative_to(out.parent.parent.parent)}")
        total_apis += api_count

    return total_apis


def extract_v100(sdk_root: Path, output_base: Path):
    """提取 SDK 1.0.0（API接口式样书/ 按模块分文件）"""
    xlsx_dir = sdk_root / "API接口式样书"
    if not xlsx_dir.exists():
        print(f"  ✗ 未找到 1.0.0 目录：{xlsx_dir}")
        return

    output_dir = output_base / "v1.0.0" / "api-spec"
    output_dir.mkdir(parents=True, exist_ok=True)

    xlsx_files = [
        f for f in sorted(xlsx_dir.glob("*.xlsx"))
        if not f.name.startswith("~$")
        and "副本" not in f.name
        and "formatted" not in f.name
        and "merge" not in f.name
        and "_API仕様書.xlsx" not in f.name  # 跳过合并总表
    ]

    print(f"\n── SDK 1.0.0（{len(xlsx_files)} 个模块文件 → {output_dir}）──")
    total = 0
    for xlsx in xlsx_files:
        name = xlsx.stem
        match = re.search(r'仕様書_(.+)$', name)
        module = match.group(1) if match else name
        try:
            total += process_xlsx_single_module(xlsx, module, output_dir, "1.0.0")
        except Exception as e:
            print(f"    ✗ {xlsx.name} 处理失败：{e}")

    print(f"  共提取 {total} 个 API")


def extract_v102(sdk_root: Path, output_base: Path):
    """提取 SDK 1.0.2（单文件全模块，使用 ecxecl_merge 最终合并版）"""
    # ecxecl_merge 是 1.0.2 的最终合并版（2026-04-09，最新）
    xlsx_path = sdk_root / "1.0.2/2.APIインターフェース仕様書/ecxecl_merge/SHIMANO Mobile SDK_API仕様書.xlsx"
    if not xlsx_path.exists():
        print(f"  ✗ 未找到 1.0.2 文件：{xlsx_path}")
        return

    output_dir = output_base / "v1.0.2" / "api-spec"
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n── SDK 1.0.2（单文件全模块 → {output_dir}）──")
    try:
        total = process_xlsx_all_modules(xlsx_path, output_dir, "1.0.2")
        print(f"  共提取 {total} 个 API")
    except Exception as e:
        print(f"  ✗ 处理失败：{e}")
        import traceback; traceback.print_exc()


# ── 入口 ────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Shimano SDK xlsx → Markdown 转换器")
    parser.add_argument("--sdk-root", default=None,
                        help="SDK 根目录（默认：~/Desktop/资料/Shimano-资料/SDK）")
    parser.add_argument("--output", default=None,
                        help="输出目录（默认：./specs）")
    parser.add_argument("--version", choices=["1.0.0", "1.0.2", "all"],
                        default="all", help="提取版本（默认：all）")
    args = parser.parse_args()

    sdk_root = Path(args.sdk_root) if args.sdk_root else \
               Path.home() / "Desktop/资料/Shimano-资料/SDK"
    output_base = Path(args.output) if args.output else \
                  Path(__file__).parent.parent / "specs"

    if not sdk_root.exists():
        print(f"错误：SDK 根目录不存在：{sdk_root}")
        sys.exit(1)

    print(f"SDK 根目录：{sdk_root}")
    print(f"输出根目录：{output_base}")
    print(f"提取版本：{args.version}")

    if args.version in ("1.0.0", "all"):
        extract_v100(sdk_root, output_base)

    if args.version in ("1.0.2", "all"):
        extract_v102(sdk_root, output_base)

    print("\n完成！")
    print(f"  specs/v1.0.0/api-spec/  ← SDK 1.0.0 接口规格")
    print(f"  specs/v1.0.2/api-spec/  ← SDK 1.0.2 接口规格")


if __name__ == "__main__":
    main()
