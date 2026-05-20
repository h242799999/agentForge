#!/usr/bin/env python3
"""
testcase_gen.py — テスト設計書 Excel 生成スクリプト

Usage:
  python3 testcase_gen.py --input testcases/.testcase_tmp.json --output testcases/output.xlsx
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl が必要です: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ── カラム定義 ──────────────────────────────────────────────────────────────
COLUMNS = [
    ("仕向け",       12),
    ("実施環境",     14),
    ("要件No",       28),
    ("大分類\nLEVEL1", 22),
    ("中分類\nLEVEL2", 22),
    ("小分類\nLEVEL3", 36),
    ("小分類\nLEVEL4", 36),
    ("小分類\nLEVEL5", 20),
    ("小分類\nLEVEL6", 20),
    ("小分類\nLEVEL7", 20),
    ("理由（説明）\nReason (description)", 24),
    ("担保内容\nCollateral content", 40),
]

JSON_KEYS = [
    "仕向け", "実施環境", "要件No",
    "大分類", "中分類",
    "小分類L3", "小分類L4", "小分類L5", "小分類L6", "小分類L7",
    "理由", "担保内容",
]

# ── スタイル定義 ────────────────────────────────────────────────────────────
THIN = Side(style="thin", color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")   # 濃い青
HEADER_FONT  = Font(name="Meiryo", bold=True, color="FFFFFF", size=9)
TITLE_FILL   = PatternFill("solid", fgColor="2E75B6")   # 中青
TITLE_FONT   = Font(name="Meiryo", bold=True, color="FFFFFF", size=11)
DATA_FONT    = Font(name="Meiryo", size=9)
WRAP_ALIGN   = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_border(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = THIN_BORDER


def _set_col_widths(ws, start_col=2):
    for i, (_, width) in enumerate(COLUMNS):
        ws.column_dimensions[get_column_letter(start_col + i)].width = width


def _write_title(ws, feature, req_no):
    """行1-2: タイトル"""
    title = f"テスト設計書 — {feature}"
    if req_no:
        title += f"  [{req_no}]"
    ws.merge_cells(start_row=1, start_column=2,
                   end_row=2, end_column=2 + len(COLUMNS) - 1)
    cell = ws.cell(row=1, column=2, value=title)
    cell.font  = TITLE_FONT
    cell.fill  = TITLE_FILL
    cell.alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 6


def _write_header(ws, header_row=3):
    """ヘッダー行"""
    ws.row_dimensions[header_row].height = 36
    for i, (label, _) in enumerate(COLUMNS):
        col = 2 + i
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border    = THIN_BORDER


def _write_data(ws, testcases, start_row=4):
    """データ行"""
    for r_offset, tc in enumerate(testcases):
        row = start_row + r_offset
        ws.row_dimensions[row].height = 30
        for c_offset, key in enumerate(JSON_KEYS):
            col  = 2 + c_offset
            val  = tc.get(key, "") or ""
            cell = ws.cell(row=row, column=col, value=val)
            cell.font      = DATA_FONT
            cell.alignment = WRAP_ALIGN
            cell.border    = THIN_BORDER
    return start_row + len(testcases) - 1


def _write_footer(ws, max_row, total):
    """フッター: 件数サマリー"""
    footer_row = max_row + 2
    cell = ws.cell(row=footer_row, column=2,
                   value=f"合計テストケース数：{total} 件")
    cell.font = Font(name="Meiryo", italic=True, size=9, color="595959")


def generate(input_path: str, output_path: str) -> int:
    """Excel を生成して保存。戻り値 = テストケース数"""
    with open(input_path, encoding="utf-8") as f:
        data = json.load(f)

    feature   = data.get("feature", "テスト設計書")
    req_no    = data.get("req_no", "")
    testcases = data.get("testcases", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "テスト設計書"

    # 列 A を非表示（余白用）
    ws.column_dimensions["A"].width = 2

    _set_col_widths(ws, start_col=2)
    _write_title(ws, feature, req_no)
    _write_header(ws, header_row=3)
    max_row = _write_data(ws, testcases, start_row=4)
    _write_footer(ws, max_row, len(testcases))

    # ウィンドウ枠固定（ヘッダー行を固定）
    ws.freeze_panes = "B4"

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return len(testcases)


def main():
    parser = argparse.ArgumentParser(description="テスト設計書 Excel 生成")
    parser.add_argument("--input",  required=True, help="JSON 中間ファイルパス")
    parser.add_argument("--output", required=True, help="出力 xlsx パス")
    args = parser.parse_args()

    if not Path(args.input).exists():
        print(f"❌ 入力ファイルが存在しません: {args.input}", file=sys.stderr)
        sys.exit(1)

    print("📊 Excel 生成中...")
    try:
        count = generate(args.input, args.output)
        print(f"✅ テスト設計書保存完了：{args.output}（計 {count} 件）")
    except Exception as e:
        print(f"❌ Excel 生成失敗：{e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
